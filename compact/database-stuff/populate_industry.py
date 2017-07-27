import sqlite3, os
from openpyxl import load_workbook
import itertools

def create_connection(db_file):
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Exception as e:
        print(e)

def create_industry_row(conn, industry_row):
    sql = ''' INSERT INTO industry (naics, bea_sector_id, description) 
                VALUES (?,?,?) '''
    cur = conn.cursor()
    cur.execute(sql, industry_row)
    return cur.lastrowid

def insert_industry_info():
    wb = load_workbook(filename='FinalNAICStoIOMatch.xlsx')
    ws = wb['NAICS to Sector']
    row_range = ws[2:465]
    for rows in row_range:
        row = [rows[3].value,rows[2].value,rows[1].value]
        with conn:
           create_industry_row(conn,row)
        
list_of_rows = []
base_dir = os.path.dirname(os.path.abspath(__file__))
database = os.path.join(os.path.dirname(base_dir), 'db.sqlite3')
conn = create_connection(database)

insert_industry_info()

